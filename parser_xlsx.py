import csv
from datetime import datetime, timedelta, time as dt_time
from openpyxl import load_workbook


def excel_serial_to_date(serial):
    if serial is None:
        return None
    if isinstance(serial, datetime):
        return serial.strftime("%d/%m/%Y")
    try:
        return (datetime(1899, 12, 30) + timedelta(days=int(serial))).strftime("%d/%m/%Y")
    except Exception:
        return str(serial)


def safe_float(v, decimals=2):
    if v is None:
        return None
    try:
        return round(float(str(v).strip()), decimals)
    except (ValueError, TypeError):
        return None


def parse_hours(value):
    if value is None:
        return None
    s = str(value).strip()
    if not s or ":" not in s:
        return None
    try:
        parts = s.split(":")
        return round(int(parts[0].strip()) + int(parts[1].strip()) / 60, 4)
    except Exception:
        return None


def parse_catalog(filepath):
    """Read the fleet catalog CSV → {frota_int: {atividade, especialidade}}."""
    catalog = {}
    with open(filepath, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter=";")
        for row in reader:
            try:
                frota = int(str(row.get("CodFrota", "")).strip())
            except ValueError:
                continue
            desc = str(row.get("descricao_especialidade", "")).strip()
            if " - " in desc:
                atividade, especialidade = desc.split(" - ", 1)
            else:
                atividade, especialidade = desc, ""
            catalog[frota] = {
                "atividade": atividade.strip(),
                "especialidade": especialidade.strip(),
                "descricao_especialidade": desc,
            }
    return catalog


def parse_report(filepath):
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    client = str(rows[0][1]).strip() if rows[0][1] else ""
    period_start = excel_serial_to_date(rows[2][1])
    period_end = excel_serial_to_date(rows[2][3])

    vehicles = []
    for row in rows[4:]:
        frota = row[0]
        desc = row[1]
        if frota is None or str(frota).strip() == "":
            break

        vehicles.append({
            "frota": int(frota),
            "descricao": str(desc).strip() if desc else "",
            "horas_periodo": parse_hours(row[2]),
            "horas_oficina": parse_hours(row[3]),
            "horas_apontadas": parse_hours(row[4]),
            "horas_disponiveis": parse_hours(row[5]),
            "disponibilidade": safe_float(row[6], 4),
            "horas_improdutivas": parse_hours(row[8]),
            "qtde_corretivas": int(row[9]) if row[9] is not None else 0,
            "mtbf_h": parse_hours(row[10]),
            "mtbf_dias": safe_float(row[11]),
            "mttr_h": parse_hours(row[12]),
            "mttr_dias": safe_float(row[13]),
        })

    with_disp = [v for v in vehicles if v["disponibilidade"] is not None]
    with_mtbf = [v for v in vehicles if v["mtbf_h"] is not None]
    with_mttr = [v for v in vehicles if v["mttr_h"] is not None]

    avg_disp = round(sum(v["disponibilidade"] for v in with_disp) / len(with_disp), 2) if with_disp else 0
    avg_mtbf = round(sum(v["mtbf_h"] for v in with_mtbf) / len(with_mtbf), 2) if with_mtbf else 0
    avg_mttr = round(sum(v["mttr_h"] for v in with_mttr) / len(with_mttr), 2) if with_mttr else 0

    return {
        "cliente": client,
        "periodo_inicio": period_start,
        "periodo_fim": period_end,
        "total_veiculos": len(vehicles),
        "summary": {
            "avg_disponibilidade": avg_disp,
            "avg_mtbf_h": avg_mtbf,
            "avg_mttr_h": avg_mttr,
            "total_corretivas": sum(v["qtde_corretivas"] for v in vehicles),
            "acima_95pct": sum(1 for v in with_disp if v["disponibilidade"] >= 95),
            "abaixo_90pct": sum(1 for v in with_disp if v["disponibilidade"] < 90),
            "com_mtbf": len(with_mtbf),
            "com_mttr": len(with_mttr),
        },
        "vehicles": vehicles,
    }


def _parse_os_datetime(date_val, time_val):
    """Parse date + time cell values from OS report → datetime or None."""
    if date_val is None:
        return None
    if isinstance(date_val, datetime):
        d = date_val.replace(second=0, microsecond=0)
    elif isinstance(date_val, (int, float)):
        d = datetime(1899, 12, 30) + timedelta(days=int(date_val))
    else:
        s = str(date_val).strip()
        if not s:
            return None
        parsed = None
        for fmt in ("%d/%m/%y", "%d/%m/%Y", "%Y-%m-%d"):
            try:
                parsed = datetime.strptime(s, fmt)
                break
            except ValueError:
                continue
        if not parsed:
            return None
        d = parsed

    if time_val is not None:
        if isinstance(time_val, dt_time):
            return d.replace(hour=time_val.hour, minute=time_val.minute, second=0, microsecond=0)
        if isinstance(time_val, datetime):
            return d.replace(hour=time_val.hour, minute=time_val.minute, second=0, microsecond=0)
        ts = str(time_val).strip()
        if ":" in ts:
            parts = ts.split(":")
            try:
                return d.replace(hour=int(parts[0]) % 24, minute=int(parts[1]), second=0, microsecond=0)
            except (ValueError, IndexError):
                pass
    return d


def _fmt_tempo(minutos):
    if minutos is None:
        return None
    m = abs(int(minutos))
    days, rem = divmod(m, 1440)
    h, mins = divmod(rem, 60)
    if days:
        return f"{days}d {h:02d}:{mins:02d}"
    return f"{h:02d}:{mins:02d}"


def parse_os_report(filepath):
    """Parse OS report xlsx (header on row 19) → list of OS dicts."""
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=19, max_row=19, values_only=True))
    col_map = {}
    for i, h in enumerate(header_row):
        if h is not None:
            key = str(h).strip()
            if key and key not in col_map:
                col_map[key] = i

    def get(row, name, default=None):
        idx = col_map.get(name)
        if idx is None or idx >= len(row):
            return default
        v = row[idx]
        return v if v is not None else default

    records = []
    for row in ws.iter_rows(min_row=20, values_only=True):
        nro_raw = get(row, "Nro OS")
        if nro_raw is None or str(nro_raw).strip() == "":
            continue
        frota_raw = get(row, "Veículo")
        try:
            frota = int(str(frota_raw).strip())
        except (ValueError, TypeError):
            continue

        status = str(get(row, "Status", "")).strip().upper()

        dt_ab  = _parse_os_datetime(get(row, "Data"),          get(row, "Hora"))
        dt_lib = _parse_os_datetime(get(row, "Data Liberação"), get(row, "Hora Liberação"))
        dt_can = _parse_os_datetime(get(row, "Data de Cancelamento"), get(row, "Hora de Cancelamento"))

        dt_fim = None
        tempo_min = None
        if status == "L" and dt_lib and dt_ab:
            dt_fim    = dt_lib
            tempo_min = int((dt_lib - dt_ab).total_seconds() / 60)
        elif status == "C" and dt_can and dt_ab:
            dt_fim    = dt_can
            tempo_min = int((dt_can - dt_ab).total_seconds() / 60)
        elif status == "A" and dt_ab:
            tempo_min = int((datetime.now() - dt_ab).total_seconds() / 60)

        records.append({
            "nro_os":        str(nro_raw).strip(),
            "status":        status,
            "frota":         frota,
            "desc_veiculo":  str(get(row, "Descrição Veículo", "")).strip(),
            "desc_problema": str(get(row, "Descrição do Problema", "")).strip(),
            "abertura_iso":  dt_ab.isoformat() if dt_ab else None,
            "abertura_str":  dt_ab.strftime("%d/%m/%y %H:%M") if dt_ab else "—",
            "fechamento_str": (dt_fim.strftime("%d/%m/%y %H:%M") if dt_fim
                               else ("Em aberto" if status == "A" else "—")),
            "tempo_str":     _fmt_tempo(tempo_min) if tempo_min is not None else "—",
            "tempo_min":     tempo_min,
        })

    return records


def build_grouped(catalog, report):
    """
    Merge catalog + report metrics, returning vehicles enriched with
    atividade/especialidade, plus an aggregate summary per specialty group.
    """
    metrics_map = {v["frota"]: v for v in report["vehicles"]}

    enriched = []
    for frota, cat in catalog.items():
        m = metrics_map.get(frota, {})
        enriched.append({
            "frota": frota,
            "atividade": cat["atividade"],
            "especialidade": cat["especialidade"],
            "descricao_especialidade": cat["descricao_especialidade"],
            "descricao": m.get("descricao", ""),
            "disponibilidade": m.get("disponibilidade"),
            "horas_periodo": m.get("horas_periodo"),
            "horas_oficina": m.get("horas_oficina"),
            "horas_improdutivas": m.get("horas_improdutivas"),
            "qtde_corretivas": m.get("qtde_corretivas", 0),
            "mtbf_h": m.get("mtbf_h"),
            "mttr_h": m.get("mttr_h"),
            "in_report": frota in metrics_map,
        })

    # Aggregate by specialty group
    groups = {}
    for v in enriched:
        key = v["descricao_especialidade"]
        if key not in groups:
            groups[key] = {
                "atividade": v["atividade"],
                "especialidade": v["especialidade"],
                "descricao_especialidade": key,
                "frotas": [],
            }
        groups[key]["frotas"].append(v)

    # Compute group-level averages
    group_list = []
    for g in groups.values():
        frotas_with_disp = [f for f in g["frotas"] if f["disponibilidade"] is not None]
        frotas_with_mtbf = [f for f in g["frotas"] if f["mtbf_h"] is not None]
        frotas_with_mttr = [f for f in g["frotas"] if f["mttr_h"] is not None]

        g["total_frotas"] = len(g["frotas"])
        g["frotas_no_report"] = sum(1 for f in g["frotas"] if f["in_report"])
        g["avg_disponibilidade"] = (
            round(sum(f["disponibilidade"] for f in frotas_with_disp) / len(frotas_with_disp), 2)
            if frotas_with_disp else None
        )
        g["avg_mtbf_h"] = (
            round(sum(f["mtbf_h"] for f in frotas_with_mtbf) / len(frotas_with_mtbf), 2)
            if frotas_with_mtbf else None
        )
        g["avg_mttr_h"] = (
            round(sum(f["mttr_h"] for f in frotas_with_mttr) / len(frotas_with_mttr), 2)
            if frotas_with_mttr else None
        )
        g["total_corretivas"] = sum(f["qtde_corretivas"] for f in g["frotas"])
        group_list.append(g)

    # Sort by atividade then especialidade
    group_list.sort(key=lambda g: (g["atividade"], g["especialidade"]))

    # Overall availability for this period (from enriched vehicles that have data)
    all_with_disp = [v for v in enriched if v["disponibilidade"] is not None]
    overall_disp = (
        round(sum(v["disponibilidade"] for v in all_with_disp) / len(all_with_disp), 2)
        if all_with_disp else None
    )

    return {
        "periodo_inicio": report["periodo_inicio"],
        "periodo_fim": report["periodo_fim"],
        "overall_disponibilidade": overall_disp,
        "groups": group_list,
        "vehicles": enriched,
    }
